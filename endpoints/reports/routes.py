from app_logger import logger
from config import DB_HOST, DB_NAME, DB_USER, DB_PASSWORD, TEMP_DIR, TEMP_NAME_LENGTH
from ..users.utils import get_current_user
from .utils import doc_to_base64
from fastapi import APIRouter, status, HTTPException, Security
import pymysql.cursors
import traceback
import openpyxl as xl
from openpyxl.styles.borders import Border, Side
from pathlib import Path
import random
import string

router = APIRouter(
    prefix='/reports',
    tags=['Report'],
    responses={404: {'detail': 'Not found'}},
)

thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


@router.get('/production', status_code=status.HTTP_200_OK)
async def get_production_report(current_user=Security(get_current_user, scopes=['production:read']),
                                start_date='2000-01-01', end_date='2099-12-31'):
    try:
        connection = pymysql.connect(host=DB_HOST,
                                     user=DB_USER,
                                     password=DB_PASSWORD,
                                     database=DB_NAME,
                                     cursorclass=pymysql.cursors.DictCursor)
        with connection:
            with connection.cursor() as cursor:
                try:
                    sql = """SELECT p.name AS product_name,
                                 amount,
                                 item_measure,
                                 m.name AS manufacturer_name,
                                 s.name AS storage_name,
                                 date,
                                 productions.comment AS comment
                             FROM `productions`
                             LEFT JOIN `products` p ON productions.product_id = p.id
                             LEFT JOIN `manufacturers` m on productions.manufacturer_id = m.id
                             LEFT JOIN `storages` s on productions.storage_id = s.id
                             WHERE storage_id!=0
                             AND `date` BETWEEN '{0}' AND '{1}'""".format(start_date, end_date)
                    cursor.execute(sql)
                    result = cursor.fetchall()
                    book = xl.Workbook()
                    sheet = book.active
                    result = list(result)
                    start_cell_row = 8
                    start_cell_col = 4
                    header_list = ['Продукция', 'Кол-во', 'Ед. изм.', 'Производитель', 'Склад', 'Дата производства', 'Комментарий']
                    report_tail = 'за все время'
                    if start_date != '2000-01-01' and end_date != '2099-12-31':
                        report_tail = f'за период с {start_date} по {end_date}'
                    elif start_date == '2000-01-01' and end_date != '2099-12-31':
                        report_tail = f'за период по {end_date}'
                    elif start_date != '2000-01-01' and end_date == '2099-12-31':
                        report_tail = f'за период с {start_date}'
                    sheet.cell(row=4, column=5).value = f'Отчет по производству продукции {report_tail}'
                    sheet.cell(row=start_cell_row - 1, column=start_cell_col - 1).value = f'№'
                    sheet.cell(row=start_cell_row - 1, column=start_cell_col - 1).border = thin_border
                    for index_row, row in enumerate(result):
                        for index_column, column in enumerate(dict(row).values()):
                            if index_column == 0:
                                sheet.cell(row=index_row + start_cell_row,
                                           column=index_column + start_cell_col - 1).value = index_row
                                sheet.cell(row=index_row + start_cell_row,
                                           column=index_column + start_cell_col - 1).border = thin_border
                            if index_row == 0:
                                sheet.cell(row=index_row + start_cell_row - 1,
                                           column=index_column + start_cell_col).value = header_list[index_column]
                                sheet.cell(row=index_row + start_cell_row - 1,
                                           column=index_column + start_cell_col).border = thin_border
                            sheet.cell(row=index_row + start_cell_row,
                                       column=index_column + start_cell_col).value = column
                            sheet.cell(row=index_row + start_cell_row,
                                       column=index_column + start_cell_col).border = thin_border
                    sheet.column_dimensions['C'].width = 5
                    sheet.column_dimensions['D'].width = 20
                    sheet.column_dimensions['E'].width = 10
                    sheet.column_dimensions['F'].width = 10
                    sheet.column_dimensions['G'].width = 20
                    sheet.column_dimensions['H'].width = 20
                    sheet.column_dimensions['I'].width = 20
                    sheet.column_dimensions['J'].width = 20
                    temp_file_name = ''.join(random.choices(string.ascii_lowercase +
                                                            string.digits,
                                                            k=TEMP_NAME_LENGTH)) + f'.xlsx'
                    temp_dir = Path(TEMP_DIR)
                    doc_full_name = temp_dir / temp_file_name
                    book.save(doc_full_name)
                    doc_base64 = doc_to_base64(str(doc_full_name))
                    doc_full_name.unlink()
                    return doc_base64
                except Exception as err:
                    lf = '\n'
                    logger.error(f'{traceback.format_exc().replace(lf, "")} : {str(err)}')
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                                        detail=f'{traceback.format_exc()} : {str(err)}')
    except Exception as err:
        lf = '\n'
        logger.error(f'{traceback.format_exc().replace(lf, "")} : {str(err)}')
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                            detail=f'{traceback.format_exc()} : {str(err)}')


@router.get('/orders', status_code=status.HTTP_200_OK)
async def get_order_report(current_user=Security(get_current_user, scopes=['order:read']),
                           start_date='2000-01-01', end_date='2099-12-31'):
    try:
        connection = pymysql.connect(host=DB_HOST,
                                     user=DB_USER,
                                     password=DB_PASSWORD,
                                     database=DB_NAME,
                                     cursorclass=pymysql.cursors.DictCursor)
        with connection:
            with connection.cursor() as cursor:
                try:
                    sql = """SELECT products.order_id,
                                   products.date,
                                   products.status,
                                   SUM(products.product_sum) AS order_sum,
                                   products.fio,
                                   products.email,
                                   products.phone,
                                   products.pickpoint_name,
                                   products.comment
                            FROM
                            (
                            SELECT o.id AS order_id,
                                   o.order_date AS date,
                                   o.status AS status,
                                   ROUND(oc.amount * prc.item_price, 2) AS product_sum,
                                   u.fio AS fio,
                                   u.email AS email,
                                   u.phone AS phone,
                                   p.name AS pickpoint_name,
                                   o.comment AS comment
                            FROM `orders` o
                            LEFT JOIN `users` u on o.user_id = u.id
                            LEFT JOIN `pickpoints` p on o.pickpoint_id = p.id
                            LEFT JOIN `order_contents` oc on o.id = oc.order_id
                            LEFT JOIN `products` pr ON oc.product_id = pr.id
                            LEFT JOIN `prices` prc ON oc.price_id = prc.id
                            WHERE `order_date` BETWEEN '{0}' AND '{1}'
                            ) products
                            GROUP BY products.order_id,
                                     products.date,
                                     products.status,
                                     products.fio,
                                     products.email,
                                     products.phone,
                                     products.pickpoint_name""".format(start_date, end_date)
                    cursor.execute(sql)
                    result = cursor.fetchall()
                    book = xl.Workbook()
                    sheet = book.active
                    result = list(result)
                    start_cell_row = 8
                    start_cell_col = 4
                    header_list = ['ИД', 'Дата заказа', 'Статус', 'Сумма', 'ФИО', 'e-mail', 'Телефон', 'Место вадачи', 'Комментарий']
                    report_tail = 'за все время'
                    if start_date != '2000-01-01' and end_date != '2099-12-31':
                        report_tail = f'за период с {start_date} по {end_date}'
                    elif start_date == '2000-01-01' and end_date != '2099-12-31':
                        report_tail = f'за период по {end_date}'
                    elif start_date != '2000-01-01' and end_date == '2099-12-31':
                        report_tail = f'за период с {start_date}'
                    sheet.cell(row=4, column=8).value = f'Отчет по заказам {report_tail}'
                    for index_row, row in enumerate(result):
                        for index_column, column in enumerate(dict(row).values()):
                            if index_row == 0:
                                sheet.cell(row=index_row + start_cell_row - 1,
                                           column=index_column + start_cell_col).value = header_list[index_column]
                                sheet.cell(row=index_row + start_cell_row - 1,
                                           column=index_column + start_cell_col).border = thin_border
                            sheet.cell(row=index_row + start_cell_row,
                                       column=index_column + start_cell_col).value = column
                            sheet.cell(row=index_row + start_cell_row,
                                       column=index_column + start_cell_col).border = thin_border
                    sheet.column_dimensions['D'].width = 5
                    sheet.column_dimensions['E'].width = 20
                    sheet.column_dimensions['F'].width = 20
                    sheet.column_dimensions['G'].width = 10
                    sheet.column_dimensions['H'].width = 30
                    sheet.column_dimensions['I'].width = 30
                    sheet.column_dimensions['J'].width = 20
                    sheet.column_dimensions['K'].width = 20
                    sheet.column_dimensions['L'].width = 30
                    temp_file_name = ''.join(random.choices(string.ascii_lowercase +
                                                            string.digits,
                                                            k=TEMP_NAME_LENGTH)) + f'.xlsx'
                    temp_dir = Path(TEMP_DIR)
                    doc_full_name = temp_dir / temp_file_name
                    book.save(doc_full_name)
                    doc_base64 = doc_to_base64(str(doc_full_name))
                    doc_full_name.unlink()
                    return doc_base64
                except Exception as err:
                    lf = '\n'
                    logger.error(f'{traceback.format_exc().replace(lf, "")} : {str(err)}')
                    raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                                        detail=f'{traceback.format_exc()} : {str(err)}')
    except Exception as err:
        lf = '\n'
        logger.error(f'{traceback.format_exc().replace(lf, "")} : {str(err)}')
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                            detail=f'{traceback.format_exc()} : {str(err)}')
